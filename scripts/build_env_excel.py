"""Generate a template Excel workbook for environment URLs and login entry points."""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = "/Users/ammarrahmouni/Projects/My-Projects/biz-buradayiz/Environment-URL-and-Login-Template.xlsx"

HEADERS = [
    "Environment",
    "Website URL",
    "Control panel URL",
    "Provider login URL",
    "Provider account / notes",
    "Admin account / notes",
    "Other links & notes",
]

ENVIRONMENTS = [
    "Staging",
    "Production",
]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Environments"

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF2F5496")
    body_font = Font(name="Calibri", size=11)
    thin = Side(style="thin", color="FFD0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(vertical="top", wrap_text=True)
    link_align = Alignment(vertical="top", wrap_text=True)
    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = border

    for row, env in enumerate(ENVIRONMENTS, start=2):
        ws.cell(row=row, column=1, value=env)
        for col in range(1, len(HEADERS) + 1):
            c = ws.cell(row=row, column=col)
            c.font = body_font
            c.border = border
            c.alignment = link_align if col in (2, 3, 4) else wrap

    widths = [22, 38, 38, 38, 32, 32, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    last_row = 1 + len(ENVIRONMENTS)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{last_row}"

    # Instructions sheet (second tab)
    inst = wb.create_sheet("How to use")
    inst.column_dimensions["A"].width = 100
    lines = [
        "This workbook is a template: fill in URLs and non-secret account identifiers only.",
        "Store passwords and secrets in a password manager or your team vault — not in this file.",
        "In Excel, you can use Insert > Link (or right-click) to turn cells into hyperlinks for one-click open.",
        "Each row is one environment (here: Staging and Production).",
    ]
    for r, line in enumerate(lines, start=1):
        inst.cell(row=r, column=1, value=line)
        inst.cell(row=r, column=1).font = Font(name="Calibri", size=11)
        inst.cell(row=r, column=1).alignment = wrap
    inst.row_dimensions[1].height = 36
    inst.row_dimensions[2].height = 36
    inst.row_dimensions[3].height = 36
    inst.row_dimensions[4].height = 36

    wb.save(OUT)
    print("Wrote", OUT)


if __name__ == "__main__":
    main()
